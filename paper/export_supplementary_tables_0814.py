#!/usr/bin/env python3
"""Export the 0814/V4 supplementary workbook to reviewable CSV files."""

from __future__ import annotations

import argparse
import csv
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


PAPER = Path(__file__).resolve().parent
DEFAULT_WORKBOOK = PAPER / "Supplementary_tables.xlsx"
DEFAULT_OUTPUT = PAPER / "supplementary_tables_0814"
EXPECTED_SHEETS = ["Index", *[f"Table S{i}" for i in range(1, 24)]]


def serialize(value: object) -> object:
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return value


def output_name(sheet_name: str) -> str:
    return sheet_name.replace(" ", "_") + ".csv"


def export(workbook_path: Path, output_dir: Path) -> None:
    workbook = load_workbook(
        workbook_path,
        read_only=True,
        data_only=False,
        keep_links=False,
    )
    if workbook.sheetnames != EXPECTED_SHEETS:
        raise RuntimeError(
            "Unexpected worksheet order:\n"
            f"observed={workbook.sheetnames}\nexpected={EXPECTED_SHEETS}"
        )

    output_dir.mkdir(parents=True, exist_ok=True)
    expected_outputs: set[Path] = set()

    for worksheet in workbook.worksheets:
        destination = output_dir / output_name(worksheet.title)
        expected_outputs.add(destination)
        with destination.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.writer(handle, lineterminator="\n")
            for row in worksheet.iter_rows(values_only=True):
                writer.writerow([serialize(value) for value in row])

    unexpected = {
        path for path in output_dir.glob("*.csv") if path not in expected_outputs
    }
    if unexpected:
        names = ", ".join(sorted(path.name for path in unexpected))
        raise RuntimeError(f"Unexpected CSV files in output directory: {names}")

    print(
        f"Exported {len(workbook.sheetnames)} worksheets from "
        f"{workbook_path} to {output_dir}"
    )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    export(args.workbook.resolve(), args.output_dir.resolve())


if __name__ == "__main__":
    main()
