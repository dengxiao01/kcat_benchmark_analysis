#!/usr/bin/env python3
"""Rebuild all manuscript tables from canonical row-level benchmark outputs."""

from __future__ import annotations

import argparse
import json
import math
from itertools import combinations
from pathlib import Path
from xml.etree import ElementTree
from zipfile import ZipFile

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from scipy.stats import rankdata, wilcoxon


BASE = Path(__file__).resolve().parent.parent
GENERATED_AUDIT_DIR = BASE / "paper" / "_generated_numeric_audit"
DEFAULT_OUTPUT = GENERATED_AUDIT_DIR / "numeric_audit_workbook_v1.2.0.xlsx"
TRUTH = BASE / "data" / "final" / "benchmark_ready_catpred.csv"
CONTEXT = BASE / "reports" / "tables" / "benchmark_ready_catpred_enriched_context.csv"
CATPRED_OVERLAP = BASE / "reports" / "tables" / "catpred_db_vs_our_benchmark_overlap.csv"
DEKP_RUN_REPORT = BASE / "reports" / "tables" / "dekp_public_retrained_run_report.csv"
RELEASE = BASE / "configs" / "benchmark_release.json"
TABLE_EXPORT_DIR = GENERATED_AUDIT_DIR / "tables"
SNAPSHOT = BASE / "paper" / "paper_statistics_v1.2.0.json"
AUDIT_TABLE_NAMES = [
    "S16_Label_audit",
    "S17_Sensitivity_subsets",
    "S18_Cluster_bootstrap",
    "S19_Cluster_wilcoxon",
    "S20_Training_overlap",
    "S21_Measurement_dispersion",
    "S22_Mutation_status",
    "S23_Substrate_direction",
    "S24_PreTKcat_variants",
]

BENCHMARK_N = len(pd.read_csv(TRUTH, usecols=["entry_id"]))

BOOTSTRAP_REPLICATES = 2000
BOOTSTRAP_SEED = 20260710

METHOD_FILES = {
    "KcatNet": BASE / "data" / "final" / "kcatnet" / "kcatnet_kcat_predictions_evaluated.csv",
    "CataPro": BASE / "data" / "final" / "catapro" / "catapro_kcat_predictions_evaluated.csv",
    "PreTKcat": BASE / "data" / "final" / "pretkcat" / "pretkcat_kcat_predictions_evaluated.csv",
    "UniKP": BASE / "data" / "final" / "unikp" / "unikp_kcat_predictions_evaluated.csv",
    "SELFprot": BASE / "data" / "final" / "selfprot" / "selfprot_kcat_predictions_evaluated.csv",
    "DLKcat": BASE / "data" / "final" / "dlkcat" / "dlkcat_kcat_predictions_evaluated.csv",
    "TurNuP": BASE / "data" / "final" / "turnup" / "turnup_kcat_predictions_evaluated.csv",
    "PMAK": BASE / "data" / "final" / "pmak" / "pmak_kcat_predictions_evaluated.csv",
    "KinForm-L": BASE / "data" / "final" / "kinform" / "kinform_kcat_predictions_evaluated.csv",
    "CatPred": BASE / "data" / "final" / "catpred" / "catpred_kcat_predictions_evaluated.csv",
    "DEKP-public-retrained": BASE
    / "data"
    / "final"
    / "dekp"
    / "dekp_public_retrained_kcat_predictions_evaluated.csv",
    "GO-HKP": BASE / "data" / "final" / "go_hkp" / "go_hkp_kcat_predictions_evaluated.csv",
}
METHOD_ORDER = list(METHOD_FILES)
NEAR_FULL = ["KcatNet", "CataPro", "UniKP", "SELFprot", "DLKcat"]
REACTION_METHODS = ["KcatNet", "TurNuP", "PMAK"]

REGIMES = {
    "KcatNet": "Released sequence+substrate checkpoint",
    "CataPro": "Released sequence+substrate checkpoint",
    "PreTKcat": "Temperature-conditioned public retraining",
    "UniKP": "Released sequence+substrate checkpoint",
    "SELFprot": "Released sequence+substrate checkpoint",
    "DLKcat": "Released sequence+substrate checkpoint",
    "TurNuP": "Reaction-aware checkpoint",
    "PMAK": "Reaction-aware checkpoint",
    "KinForm-L": "Method-specific checkpoint subset",
    "CatPred": "Method-specific checkpoint subset",
    "DEKP-public-retrained": "Structure-aware public retraining",
    "GO-HKP": "Functional-assignment baseline",
}
MODALITIES = {
    "KcatNet": "sequence + substrate SMILES",
    "CataPro": "sequence + substrate SMILES",
    "PreTKcat": "sequence + substrate SMILES + temperature",
    "UniKP": "sequence + substrate SMILES",
    "SELFprot": "sequence + substrate SMILES",
    "DLKcat": "sequence + substrate SMILES",
    "TurNuP": "reaction + enzyme",
    "PMAK": "reaction + enzyme",
    "KinForm-L": "sequence + substrate SMILES + released features",
    "CatPred": "sequence + substrate SMILES",
    "DEKP-public-retrained": "sequence + substrate SMILES + structure",
    "GO-HKP": "GO hierarchy + functional assignment",
}
PREDICTION_COLUMNS = {
    method: (
        "DEKP_public_retrained_predicted_kcat_s^-1"
        if method == "DEKP-public-retrained"
        else f"{method}_predicted_kcat_s^-1"
    )
    for method in METHOD_ORDER
}

METHOD_INPUT_SPECS = {
    "KcatNet": {
        "path": BASE / "data/final/kcatnet/kcatnet_kcat_input_metadata.csv",
        "sequence_column": "kcatnet_model_sequence",
        "sequence_policy": "first 500 + last 500 aa when length >1,000; otherwise full sequence",
    },
    "CataPro": {
        "path": BASE / "data/final/catapro/catapro_kcat_input_metadata.csv",
        "sequence_column": "sequence",
        "sequence_policy": "full benchmark sequence supplied to adapter",
    },
    "PreTKcat": {
        "path": BASE / "data/final/pretkcat/pretkcat_kcat_input_metadata.csv",
        "sequence_column": "pretkcat_model_sequence",
        "sequence_policy": "first 500 + last 500 aa when length >1,000; otherwise full sequence",
    },
    "UniKP": {
        "path": BASE / "data/final/unikp/unikp_kcat_input_metadata.csv",
        "sequence_column": "unikp_model_sequence",
        "sequence_policy": "first 500 + last 500 aa when length >1,000; otherwise full sequence",
    },
    "SELFprot": {
        "path": BASE / "data/final/selfprot/selfprot_kcat_input.csv",
        "sequence_column": "sequence",
        "sequence_policy": "full benchmark sequence supplied to adapter",
    },
    "DLKcat": {
        "path": BASE / "data/final/dlkcat/dlkcat_kcat_input_metadata.csv",
        "sequence_column": "sequence",
        "sequence_policy": "full benchmark sequence supplied to adapter",
    },
    "TurNuP": {
        "path": BASE / "data/final/turnup/turnup_kcat_input.csv",
        "sequence_column": "sequence",
        "reaction_column": "reaction_smiles",
        "sequence_policy": "full benchmark sequence on reaction-complete rows",
    },
    "PMAK": {
        "path": BASE / "data/final/pmak/pmak_kcat_input.csv",
        "sequence_column": "sequence",
        "reaction_column": "reaction_smiles",
        "sequence_policy": "full benchmark sequence on reaction-complete rows",
    },
    "KinForm-L": {
        "path": BASE / "data/final/kinform/kinform_kcat_input_metadata.csv",
        "sequence_column": "kinform_model_sequence",
        "sequence_policy": "first 749 + last 749 aa when length >1,499; otherwise full sequence",
    },
    "CatPred": {
        "path": BASE / "data/final/catpred/catpred_kcat_input.csv",
        "sequence_column": "sequence",
        "sequence_policy": "full benchmark sequence supplied to adapter",
    },
    "DEKP-public-retrained": {
        "path": BASE / "data/final/dekp/dekp_kcat_input_metadata.csv",
        "sequence_column": "sequence",
        "sequence_policy": "full benchmark sequence supplied with substrate and structure features",
    },
    "GO-HKP": {
        "path": BASE / "data/final/go_hkp/go_hkp_kcat_input_metadata.csv",
        "sequence_column": None,
        "sequence_policy": "no direct sequence input; gene/UniProt GO assignment is used",
    },
}

REACTION_INPUT_ROLES = {
    method: "model equation is provenance only; complete reaction is not an inference input"
    for method in METHOD_ORDER
}
REACTION_INPUT_ROLES.update(
    {
        "TurNuP": "complete reaction SMILES is an inference input",
        "PMAK": "complete reaction SMILES is an inference input",
        "GO-HKP": "reaction/gene identifiers support functional assignment; equation is provenance",
    }
)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--bootstrap-replicates", type=int, default=BOOTSTRAP_REPLICATES)
    parser.add_argument("--seed", type=int, default=BOOTSTRAP_SEED)
    return parser.parse_args()


def bool_series(series: pd.Series) -> pd.Series:
    if pd.api.types.is_bool_dtype(series):
        return series.fillna(False)
    return series.fillna("").astype(str).str.lower().isin({"true", "1", "yes"})


def load_methods() -> dict[str, pd.DataFrame]:
    methods = {}
    for method, path in METHOD_FILES.items():
        frame = pd.read_csv(path)
        required = {"entry_id", "true_kcat_log10", "prediction_log10", "prediction_kcat"}
        missing = required.difference(frame.columns)
        if missing:
            raise ValueError(f"{method} is missing columns: {sorted(missing)}")
        if frame["entry_id"].duplicated().any():
            raise ValueError(f"{method} contains duplicate entry_id values")
        numeric = frame[["true_kcat_log10", "prediction_log10", "prediction_kcat"]].apply(
            pd.to_numeric, errors="coerce"
        )
        if not np.isfinite(numeric).all().all():
            raise ValueError(f"{method} contains non-finite truth or prediction values")
        expected_linear = np.power(10.0, numeric["prediction_log10"])
        if not np.allclose(expected_linear, numeric["prediction_kcat"], rtol=1e-8, atol=1e-8):
            raise ValueError(f"{method} has inconsistent linear and log10 predictions")
        methods[method] = frame.copy()
    return methods


def load_audit_tables() -> dict[str, pd.DataFrame]:
    tables = {}
    for name in AUDIT_TABLE_NAMES:
        path = TABLE_EXPORT_DIR / f"{name}.csv"
        if not path.exists():
            raise FileNotFoundError(
                f"Missing {path}; run paper/build_submission_audits.py before rebuilding tables"
            )
        tables[name] = pd.read_csv(path)
    return tables


def load_method_inputs() -> dict[str, pd.DataFrame]:
    inputs = {}
    for method, specification in METHOD_INPUT_SPECS.items():
        path = Path(specification["path"])
        frame = pd.read_csv(path, low_memory=False)
        if "entry_id" not in frame.columns:
            raise ValueError(f"{method} input metadata has no entry_id: {path}")
        if frame["entry_id"].duplicated().any():
            raise ValueError(f"{method} input metadata contains duplicate entry_id values")

        selected = pd.DataFrame({"entry_id": frame["entry_id"].astype(str)})
        sequence_column = specification.get("sequence_column")
        if sequence_column is None:
            selected["method_sequence_input"] = pd.NA
        else:
            if sequence_column not in frame.columns:
                raise ValueError(f"{method} input metadata lacks {sequence_column}: {path}")
            selected["method_sequence_input"] = frame[sequence_column].fillna("").astype(str)

        reaction_column = specification.get("reaction_column")
        if reaction_column is None:
            selected["method_reaction_input"] = pd.NA
        else:
            if reaction_column not in frame.columns:
                raise ValueError(f"{method} input metadata lacks {reaction_column}: {path}")
            selected["method_reaction_input"] = frame[reaction_column].fillna("").astype(str)
        inputs[method] = selected
    return inputs


def format_stoichiometry(value: float) -> str:
    magnitude = abs(float(value))
    if np.isclose(magnitude, 1.0):
        return ""
    if np.isclose(magnitude, round(magnitude)):
        return f"{int(round(magnitude))} "
    return f"{magnitude:g} "


def clean_model_metabolite_name(value: object) -> str:
    text = str(value).strip()
    tokens = text.split()
    midpoint = len(tokens) // 2
    if tokens and len(tokens) % 2 == 0 and tokens[:midpoint] == tokens[midpoint:]:
        return " ".join(tokens[:midpoint])
    return text


def reaction_equation(
    reactants: list[tuple[float, str]],
    products: list[tuple[float, str]],
    names: dict[str, str] | None = None,
) -> str:
    def side(terms: list[tuple[float, str]]) -> str:
        formatted = []
        for coefficient, metabolite_id in terms:
            label = names.get(metabolite_id, metabolite_id) if names is not None else metabolite_id
            formatted.append(f"{format_stoichiometry(coefficient)}{label}")
        return " + ".join(formatted) if formatted else "0"

    return f"{side(reactants)} -> {side(products)}"


def build_reaction_provenance() -> pd.DataFrame:
    rows = []

    ecoli_model = json.loads((BASE / "eciML1515.json").read_text(encoding="utf-8"))
    ecoli_names = {
        str(metabolite["id"]): clean_model_metabolite_name(
            metabolite.get("name") or metabolite["id"]
        )
        for metabolite in ecoli_model["metabolites"]
    }
    for reaction in ecoli_model["reactions"]:
        reactants = [
            (coefficient, metabolite_id)
            for metabolite_id, coefficient in reaction["metabolites"].items()
            if float(coefficient) < 0
        ]
        products = [
            (coefficient, metabolite_id)
            for metabolite_id, coefficient in reaction["metabolites"].items()
            if float(coefficient) > 0
        ]
        rows.append(
            {
                "species": "ecoli",
                "reaction_id": str(reaction["id"]),
                "reaction_name": str(reaction.get("name") or reaction["id"]),
                "reaction_equation_ids_model_forward": reaction_equation(reactants, products),
                "reaction_equation_names_model_forward": reaction_equation(
                    reactants, products, ecoli_names
                ),
                "reaction_reversible_in_model": float(reaction.get("lower_bound", 0)) < 0,
            }
        )

    yeast_root = ElementTree.parse(BASE / "yeast-GEM.xml").getroot()
    core_namespace = yeast_root.tag.split("}", 1)[0].lstrip("{")
    namespace = {"sbml": core_namespace}
    yeast_names = {
        str(species.attrib["id"]): clean_model_metabolite_name(
            species.attrib.get("name") or species.attrib["id"]
        )
        for species in yeast_root.findall(".//sbml:listOfSpecies/sbml:species", namespace)
    }
    for reaction in yeast_root.findall(".//sbml:listOfReactions/sbml:reaction", namespace):
        reactants_element = reaction.find("sbml:listOfReactants", namespace)
        products_element = reaction.find("sbml:listOfProducts", namespace)
        reactants = [] if reactants_element is None else [
            (float(reference.attrib.get("stoichiometry", "1")), str(reference.attrib["species"]))
            for reference in reactants_element.findall("sbml:speciesReference", namespace)
        ]
        products = [] if products_element is None else [
            (float(reference.attrib.get("stoichiometry", "1")), str(reference.attrib["species"]))
            for reference in products_element.findall("sbml:speciesReference", namespace)
        ]
        rows.append(
            {
                "species": "yeast",
                "reaction_id": str(reaction.attrib["id"]),
                "reaction_name": str(reaction.attrib.get("name") or reaction.attrib["id"]),
                "reaction_equation_ids_model_forward": reaction_equation(reactants, products),
                "reaction_equation_names_model_forward": reaction_equation(
                    reactants, products, yeast_names
                ),
                "reaction_reversible_in_model": reaction.attrib.get("reversible", "false").lower()
                == "true",
            }
        )

    provenance = pd.DataFrame(rows)
    if provenance.duplicated(["species", "reaction_id"]).any():
        raise ValueError("Source models contain duplicate species/reaction identifiers")
    return provenance


def metrics(frame: pd.DataFrame) -> dict[str, float]:
    truth = pd.to_numeric(frame["true_kcat_log10"], errors="raise").to_numpy(float)
    prediction = pd.to_numeric(frame["prediction_log10"], errors="raise").to_numpy(float)
    error = prediction - truth
    ss_res = float(np.square(error).sum())
    ss_tot = float(np.square(truth - truth.mean()).sum())
    return {
        "n": len(frame),
        "mae_log10": float(np.abs(error).mean()),
        "rmse_log10": float(np.sqrt(np.square(error).mean())),
        "pearson_log10": float(np.corrcoef(truth, prediction)[0, 1]) if len(frame) > 1 else np.nan,
        "spearman_log10": float(pd.Series(truth).corr(pd.Series(prediction), method="spearman"))
        if len(frame) > 1
        else np.nan,
        "r2_log10": float(1.0 - ss_res / ss_tot) if ss_tot else np.nan,
        "bias_log10": float(error.mean()),
        "median_abs_error_log10": float(np.median(np.abs(error))),
        "within_0.3_log10_fraction": float((np.abs(error) <= 0.3).mean()),
        "within_1.0_log10_fraction": float((np.abs(error) <= 1.0).mean()),
    }


def spearman_array(truth: np.ndarray, prediction: np.ndarray) -> float:
    truth_rank = rankdata(truth)
    pred_rank = rankdata(prediction)
    return float(np.corrcoef(truth_rank, pred_rank)[0, 1])


def bootstrap_method(frame: pd.DataFrame, method: str, replicates: int, seed: int) -> pd.DataFrame:
    truth = frame["true_kcat_log10"].to_numpy(float)
    prediction = frame["prediction_log10"].to_numpy(float)
    error = prediction - truth
    n = len(frame)
    rng = np.random.default_rng(seed)
    samples = {
        "mae_log10": np.empty(replicates),
        "spearman_log10": np.empty(replicates),
        "within_0.3_fraction": np.empty(replicates),
        "within_1.0_fraction": np.empty(replicates),
    }
    for index in range(replicates):
        selected = rng.integers(0, n, size=n)
        sampled_error = error[selected]
        samples["mae_log10"][index] = np.abs(sampled_error).mean()
        samples["spearman_log10"][index] = spearman_array(truth[selected], prediction[selected])
        samples["within_0.3_fraction"][index] = (np.abs(sampled_error) <= 0.3).mean()
        samples["within_1.0_fraction"][index] = (np.abs(sampled_error) <= 1.0).mean()

    observed = metrics(frame)
    estimates = {
        "mae_log10": observed["mae_log10"],
        "spearman_log10": observed["spearman_log10"],
        "within_0.3_fraction": observed["within_0.3_log10_fraction"],
        "within_1.0_fraction": observed["within_1.0_log10_fraction"],
    }
    rows = []
    for metric_name, values in samples.items():
        low, high = np.percentile(values, [2.5, 97.5])
        rows.append(
            {
                "method": method,
                "n": n,
                "metric": metric_name,
                "estimate": estimates[metric_name],
                "bootstrap_ci_low_95": float(low),
                "bootstrap_ci_high_95": float(high),
                "bootstrap_replicates": replicates,
                "seed": seed,
            }
        )
    return pd.DataFrame(rows)


def frame_by_ids(frame: pd.DataFrame, identifiers: set[str]) -> pd.DataFrame:
    return frame[frame["entry_id"].isin(identifiers)].copy()


def build_table0_wide(truth: pd.DataFrame, methods: dict[str, pd.DataFrame]) -> pd.DataFrame:
    columns = [
        "entry_id",
        "species",
        "reaction_id",
        "gene_id",
        "uniprot_id",
        "ec_number",
        "substrate_name",
        "SMILES",
        "true_kcat",
        "true_kcat_log10",
    ]
    table = truth[columns].rename(
        columns={
            "SMILES": "substrate_smiles",
            "true_kcat": "experimental_kcat_s^-1",
            "true_kcat_log10": "experimental_log10_kcat",
        }
    )
    for method in METHOD_ORDER:
        prediction = methods[method][["entry_id", "prediction_kcat"]].rename(
            columns={"prediction_kcat": PREDICTION_COLUMNS[method]}
        )
        table = table.merge(prediction, on="entry_id", how="left", validate="one_to_one")
    return table


def build_table0(
    truth: pd.DataFrame,
    context: pd.DataFrame,
    methods: dict[str, pd.DataFrame],
    method_inputs: dict[str, pd.DataFrame],
    reaction_provenance: pd.DataFrame,
) -> pd.DataFrame:
    context_fields = context[["entry_id", "substrate_id"]].drop_duplicates("entry_id")
    base = truth[
        [
            "entry_id",
            "species",
            "gene_id",
            "uniprot_id",
            "ec_number",
            "sequence",
            "reaction_id",
            "substrate_name",
            "SMILES",
            "substrate_stoichiometry",
            "candidate_selection_policy",
            "substrate_pubchem_cid",
            "substrate_parent_inchikey_connectivity",
            "substrate_role_class",
            "substrate_role_group",
            "substrate_role_evidence",
            "substrate_role_confidence",
            "experimental_substrate_support",
            "true_kcat",
            "true_kcat_log10",
            "source_database",
            "match_level",
            "source_record_ids",
            "n_measurements",
        ]
    ].merge(context_fields, on="entry_id", how="left", validate="one_to_one")
    base = base.merge(
        reaction_provenance,
        on=["species", "reaction_id"],
        how="left",
        validate="many_to_one",
    )
    if base["reaction_equation_ids_model_forward"].isna().any():
        missing = base.loc[
            base["reaction_equation_ids_model_forward"].isna(), ["species", "reaction_id"]
        ].drop_duplicates()
        raise ValueError(f"Missing source-model reaction equations: {missing.to_dict('records')}")

    base = base.rename(
        columns={
            "sequence": "benchmark_sequence",
            "substrate_id": "evaluated_metabolite_id",
            "substrate_name": "evaluated_metabolite_name",
            "SMILES": "evaluated_metabolite_smiles",
            "true_kcat": "experimental_kcat_s^-1",
            "true_kcat_log10": "experimental_log10_kcat",
            "source_database": "experimental_source_database",
            "match_level": "experimental_match_level",
            "source_record_ids": "experimental_source_record_ids",
            "n_measurements": "experimental_measurement_count",
        }
    )

    blocks = []
    for method in METHOD_ORDER:
        block = base.copy()
        block.insert(0, "method", method)
        block.insert(1, "inference_regime", REGIMES[method])
        block.insert(2, "inference_time_modality", MODALITIES[method])

        prepared = method_inputs[method].copy()
        prepared["method_input_prepared"] = True
        block = block.merge(prepared, on="entry_id", how="left", validate="one_to_one")
        prediction = methods[method][
            ["entry_id", "prediction_kcat", "prediction_log10"]
        ].rename(
            columns={
                "prediction_kcat": "predicted_kcat_s^-1",
                "prediction_log10": "predicted_log10_kcat",
            }
        )
        block = block.merge(prediction, on="entry_id", how="left", validate="one_to_one")

        predicted = block["predicted_kcat_s^-1"].notna()
        input_prepared = block["method_input_prepared"].eq(True)
        block["prediction_status"] = np.select(
            [predicted, input_prepared],
            ["predicted", "input_prepared_no_valid_prediction"],
            default="outside_method_input_scope",
        )

        if method == "GO-HKP":
            block["method_sequence_status"] = "not_a_direct_model_input"
        else:
            sequence_available = block["method_sequence_input"].notna() & block[
                "method_sequence_input"
            ].astype(str).ne("")
            same_sequence = sequence_available & block["method_sequence_input"].astype(str).eq(
                block["benchmark_sequence"].astype(str)
            )
            block["method_sequence_status"] = np.select(
                [same_sequence, sequence_available],
                ["full_sequence", "truncated_sequence"],
                default="not_prepared_outside_method_scope",
            )

        block["sequence_input_policy"] = METHOD_INPUT_SPECS[method]["sequence_policy"]
        block["reaction_input_role"] = REACTION_INPUT_ROLES[method]
        block["absolute_error_log10"] = (
            block["predicted_log10_kcat"] - block["experimental_log10_kcat"]
        ).abs()
        blocks.append(block)

    table = pd.concat(blocks, ignore_index=True)
    columns = [
        "method",
        "inference_regime",
        "inference_time_modality",
        "prediction_status",
        "entry_id",
        "species",
        "gene_id",
        "uniprot_id",
        "ec_number",
        "benchmark_sequence",
        "method_sequence_input",
        "method_sequence_status",
        "sequence_input_policy",
        "reaction_id",
        "reaction_name",
        "reaction_equation_ids_model_forward",
        "reaction_equation_names_model_forward",
        "reaction_reversible_in_model",
        "reaction_input_role",
        "method_reaction_input",
        "evaluated_metabolite_id",
        "evaluated_metabolite_name",
        "evaluated_metabolite_smiles",
        "substrate_stoichiometry",
        "candidate_selection_policy",
        "substrate_pubchem_cid",
        "substrate_parent_inchikey_connectivity",
        "substrate_role_class",
        "substrate_role_group",
        "substrate_role_evidence",
        "substrate_role_confidence",
        "experimental_substrate_support",
        "experimental_source_database",
        "experimental_match_level",
        "experimental_source_record_ids",
        "experimental_measurement_count",
        "experimental_kcat_s^-1",
        "experimental_log10_kcat",
        "predicted_kcat_s^-1",
        "predicted_log10_kcat",
        "absolute_error_log10",
    ]
    return table[columns]


def build_table1(methods: dict[str, pd.DataFrame]) -> pd.DataFrame:
    reaction_ids = set.intersection(
        *(set(methods[method]["entry_id"].astype(str)) for method in REACTION_METHODS)
    )
    reaction_n = len(reaction_ids)
    catpred_n = len(methods["CatPred"])
    kinform_n = len(methods["KinForm-L"])
    pretkcat = methods["PreTKcat"]
    temperature_imputed = int(bool_series(pretkcat["pretkcat_temperature_imputed"]).sum())
    exact_removed = int(pretkcat["pretkcat_train_rows_removed_exact_pair"].iloc[0])
    dekp_n = len(methods["DEKP-public-retrained"])
    go_n = len(methods["GO-HKP"])
    return pd.DataFrame(
        [
            {
                "Inference regime": "Released sequence+substrate checkpoints",
                "Methods": "DLKcat; UniKP; CataPro; KcatNet; SELFprot",
                "Inference-time inputs": "Enzyme sequence and one substrate representation",
                "Applicable records": f"{BENCHMARK_N}/{BENCHMARK_N} (100%)",
                "Comparison rule": f"Head-to-head comparison on {BENCHMARK_N} shared rows",
                "Primary caveat": "Training-corpus proximity differs by method and is reported separately.",
            },
            {
                "Inference regime": "Temperature-conditioned public retraining",
                "Methods": "PreTKcat",
                "Inference-time inputs": "Enzyme sequence, substrate SMILES, and measured or imputed temperature",
                "Applicable records": f"{BENCHMARK_N}/{BENCHMARK_N} (100%)",
                "Comparison rule": "Report separately from released sequence+substrate checkpoints",
                "Primary caveat": (
                    f"Locally fitted exact-excluded model; {temperature_imputed} temperatures imputed; "
                    f"{exact_removed} public training rows removed. Raw-public and near-excluded variants are in S24."
                ),
            },
            {
                "Inference regime": "Reaction-aware checkpoints",
                "Methods": "TurNuP; PMAK",
                "Inference-time inputs": "Enzyme sequence and complete reaction representation",
                "Applicable records": f"{reaction_n}/{BENCHMARK_N} ({100 * reaction_n / BENCHMARK_N:.1f}%)",
                "Comparison rule": f"Compare with KcatNet on the common {reaction_n}-row subset",
                "Primary caveat": "Requires complete reactant and product structures.",
            },
            {
                "Inference regime": "Method-specific checkpoint subsets",
                "Methods": "CatPred; KinForm-L",
                "Inference-time inputs": "Sequence and substrate plus method-specific preprocessing or released assets",
                "Applicable records": (
                    f"{catpred_n}/{BENCHMARK_N} ({100 * catpred_n / BENCHMARK_N:.1f}%); "
                    f"{kinform_n}/{BENCHMARK_N} ({100 * kinform_n / BENCHMARK_N:.1f}%)"
                ),
                "Comparison rule": "Interpret within each achieved applicability domain",
                "Primary caveat": "Different row sets; available-case counts and strictly paired comparisons are reported separately.",
            },
            {
                "Inference regime": "Structure-aware public retraining",
                "Methods": "DEKP-public-retrained",
                "Inference-time inputs": "Sequence, substrate, and protein structure graph",
                "Applicable records": f"{dekp_n}/{BENCHMARK_N} ({100 * dekp_n / BENCHMARK_N:.1f}%)",
                "Comparison rule": "Assess public reproducibility, not the unreleased author model",
                "Primary caveat": "Locally fitted reconstruction; final author kcat weights were not released.",
            },
            {
                "Inference regime": "Functional-assignment baseline",
                "Methods": "GO-HKP",
                "Inference-time inputs": "GO hierarchy and GO-linked kcat statistics",
                "Applicable records": f"{go_n}/{BENCHMARK_N} ({100 * go_n / BENCHMARK_N:.1f}%)",
                "Comparison rule": "Report as a transparent assignment baseline and by species",
                "Primary caveat": "E. coli and yeast use different GO evidence routes.",
            },
        ]
    )


def build_full_metrics(methods: dict[str, pd.DataFrame]) -> pd.DataFrame:
    rows = []
    for method in METHOD_ORDER:
        result = metrics(methods[method])
        rows.append(
            {
                "Method": method,
                "n": result["n"],
                "Coverage (%)": 100.0 * result["n"] / BENCHMARK_N,
                "MAE log10": result["mae_log10"],
                "RMSE log10": result["rmse_log10"],
                "Pearson": result["pearson_log10"],
                "Spearman": result["spearman_log10"],
                "R2 log10": result["r2_log10"],
                "Bias log10": result["bias_log10"],
                "Median absolute error log10": result["median_abs_error_log10"],
                "Within 2-fold (%)": 100.0 * result["within_0.3_log10_fraction"],
                "Within 10-fold (%)": 100.0 * result["within_1.0_log10_fraction"],
                "Inference regime": REGIMES[method],
                "Modality": MODALITIES[method],
            }
        )
    return pd.DataFrame(rows)


def build_table2(
    methods: dict[str, pd.DataFrame],
    cluster_bootstrap: pd.DataFrame,
) -> pd.DataFrame:
    """Keep methods with different applicability domains out of one flat ranking."""
    reaction_ids = set.intersection(
        *(set(methods[method]["entry_id"].astype(str)) for method in REACTION_METHODS)
    )
    reaction_n = len(reaction_ids)
    specifications = [
        (f"A. Released sequence+substrate checkpoints; all {BENCHMARK_N} records", method, None)
        for method in NEAR_FULL
    ]
    specifications += [
        (f"B. Temperature-conditioned public retraining; all {BENCHMARK_N} records", "PreTKcat", None),
        (f"C. Reaction-aware strictly paired comparison; common {reaction_n} records", "KcatNet", reaction_ids),
        (f"C. Reaction-aware strictly paired comparison; common {reaction_n} records", "TurNuP", reaction_ids),
        (f"C. Reaction-aware strictly paired comparison; common {reaction_n} records", "PMAK", reaction_ids),
        ("D. Method-specific applicability; achieved subset", "KinForm-L", None),
        ("D. Method-specific applicability; achieved subset", "CatPred", None),
        (f"E. Structure-aware public retraining; all {BENCHMARK_N} records", "DEKP-public-retrained", None),
        ("F. Functional-assignment baseline; achieved subset", "GO-HKP", None),
    ]

    rows = []
    for panel, method, identifiers in specifications:
        frame = methods[method]
        scope = "achieved_evaluation_set"
        if identifiers is not None:
            frame = frame_by_ids(frame, identifiers)
            if method == "KcatNet":
                scope = f"reaction_aware_common_{reaction_n}"
        result = metrics(frame)
        ci = cluster_bootstrap[
            cluster_bootstrap["analysis_scope"].eq(scope)
            & cluster_bootstrap["method"].eq(method)
            & cluster_bootstrap["cluster_type"].eq("pair")
        ]
        if len(ci) != 1:
            raise ValueError(
                f"Expected one pair-cluster interval for {method} in {scope}; found {len(ci)}"
            )
        interval = ci.iloc[0]
        rows.append(
            {
                "Evaluation panel and scope": panel,
                "Method": method,
                "n": int(result["n"]),
                f"Coverage of {BENCHMARK_N} (%)": round(100.0 * result["n"] / BENCHMARK_N, 1),
                "MAE (pair-cluster 95% CI)": (
                    f"{result['mae_log10']:.3f} "
                    f"({interval['cluster_bootstrap_ci_low_95']:.3f}-"
                    f"{interval['cluster_bootstrap_ci_high_95']:.3f})"
                ),
                "Spearman rho": round(float(result["spearman_log10"]), 3),
                "Bias": round(float(result["bias_log10"]), 3),
                "Within 10-fold (%)": round(
                    100.0 * float(result["within_1.0_log10_fraction"]), 1
                ),
            }
        )
    return pd.DataFrame(rows)


def build_method_details(methods: dict[str, pd.DataFrame]) -> pd.DataFrame:
    dekp_run = pd.read_csv(DEKP_RUN_REPORT).iloc[0]
    pretkcat_run = methods["PreTKcat"].iloc[0]
    pretkcat_temperature_imputed = int(
        bool_series(methods["PreTKcat"]["pretkcat_temperature_imputed"]).sum()
    )
    reaction_missing = BENCHMARK_N - len(methods["TurNuP"])
    catpred_missing = BENCHMARK_N - len(methods["CatPred"])
    kinform_missing = BENCHMARK_N - len(methods["KinForm-L"])
    go_missing = BENCHMARK_N - len(methods["GO-HKP"])
    kcatnet_truncated = int(
        bool_series(methods["KcatNet"]["kcatnet_sequence_truncated"]).sum()
    )
    details = {
        "CataPro": (
            "Enzyme sequence; substrate SMILES; variant type set to wild",
            "ProtT5/MolT5-derived features, fingerprints, and neural prediction",
            "Mean of 10 released fold-specific models",
            f"All {BENCHMARK_N} rows scored; no product-side input.",
        ),
        "DLKcat": (
            "Enzyme sequence; substrate name and SMILES",
            "Substrate molecular graph with protein sequence CNN and attention",
            "Official code and checkpoint",
            f"All {BENCHMARK_N} rows scored; no product-side input.",
        ),
        "KcatNet": (
            "Enzyme sequence; substrate SMILES",
            "Protein language-model and molecular graph features with geometric learning",
            "Official public checkpoint model_KcatNet.pt",
            f"{kcatnet_truncated} sequences were truncated to the 1,000-residue model limit.",
        ),
        "PreTKcat": (
            "Enzyme sequence; substrate SMILES; temperature",
            "ProtT5, MolGNet, and two normalized temperature features with ExtraTrees",
            (
                f"Public-data retraining: {int(pretkcat_run['pretkcat_train_rows']):,} fitted rows, "
                f"{int(pretkcat_run['pretkcat_feature_dim']):,} features, "
                f"{int(pretkcat_run['pretkcat_n_estimators'])} trees, seed 42"
            ),
            (
                "No fitted kcat regressor released; "
                f"{int(pretkcat_run['pretkcat_train_rows_removed_exact_pair']):,} standardized "
                f"sequence-parent rows excluded; {pretkcat_temperature_imputed} temperatures imputed to 30 C; "
                "raw-public and near-excluded reconstructions are reported in S24."
            ),
        ),
        "SELFprot": (
            "Enzyme sequence; substrate SMILES",
            "Modular protein/molecule transformer with joint kcat head",
            "Released fold-1 checkpoint",
            "Only fold 1 was present in the released weight package.",
        ),
        "UniKP": (
            "Enzyme sequence; substrate SMILES",
            "Mean-pooled ProtT5 and SMILES Transformer features with ExtraTrees",
            "Official public regressor",
            "Sequences over 1,000 residues use the first and last 500 residues.",
        ),
        "PMAK": (
            "Complete reaction SMILES; enzyme sequence",
            "ProtT5 and RXNFP features with residue-aware attention",
            "Mean of five released reaction-cold fold checkpoints",
            f"Requires complete reaction SMILES; {reaction_missing} benchmark rows were outside the input domain.",
        ),
        "TurNuP": (
            "Reactant SMILES; product SMILES; enzyme sequence",
            "DRFP reaction fingerprint and task-specific ESM-1b/ESP representation with gradient boosting",
            "Official public code and checkpoint",
            f"Requires complete reactant and product representations; {reaction_missing} rows were outside the input domain.",
        ),
        "CatPred": (
            "Substrate SMILES; enzyme sequence; generated protein-record cache key",
            "Production kcat ensemble using sequence attention/ESM2 and a D-MPNN substrate encoder",
            "Official production kcat checkpoint ensemble",
            f"{catpred_missing} rows did not yield valid predictions; production kcat inference did not use 3D EGNN.",
        ),
        "KinForm-L": (
            "Enzyme sequence; substrate SMILES; released precomputed protein features",
            "ESMC, ESM-2 and ProtT5 global/binding-site features, SMILES Transformer, and ExtraTrees",
            "Released KinForm-L asset bundle",
            f"Only {len(methods['KinForm-L'])} rows had all released lookup and feature assets; {kinform_missing} rows remained unavailable.",
        ),
        "DEKP-public-retrained": (
            "Enzyme sequence; substrate SMILES; AlphaFold/PDB residue graph",
            "DEKP MetaDecoder with molecular, sequence-CNN, and structure-graph features",
            (
                f"Public retraining: {int(dekp_run['public_rows_used']):,} rows; "
                f"best epoch {int(dekp_run['best_epoch'])}; seed 3407"
            ),
            (
                "Final kcat weights were not released; "
                f"{int(dekp_run['public_exact_pair_overlap_rows']):,} standardized "
                "sequence-parent pairs were excluded before fitting."
            ),
        ),
        "GO-HKP": (
            "DeepGO-SE reaction assignments or UniProt GO annotations",
            "GO hierarchy and organism-filtered GO-linked kcat medians",
            "Benchmark functional-assignment baseline",
            f"E. coli and yeast use different GO evidence routes; {go_missing} rows were unassigned; not a direct regression model.",
        ),
    }
    rows = []
    for method in METHOD_ORDER:
        inference_inputs, representation, status, caveat = details[method]
        rows.append(
            {
                "method": method,
                "inference_regime": REGIMES[method],
                "inference_inputs": inference_inputs,
                "representation_or_model": representation,
                "implementation_status": status,
                "main_caveat": caveat,
                "n": len(methods[method]),
                "coverage_percent": 100.0 * len(methods[method]) / BENCHMARK_N,
            }
        )
    return pd.DataFrame(rows)


def build_provenance(truth: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    label_map = {
        "species_ec_substrate_id": "Species + EC + substrate ID",
        "species_ec_uniprot_substrate_id": "Species + EC + UniProt + substrate ID",
        "species_ec_substrate_name": "Species + EC + substrate name",
        "species_ec_uniprot_substrate_name": "Species + EC + UniProt + substrate name",
    }
    matching = (
        truth.assign(**{"Matching level": truth["match_level"].map(label_map).fillna(truth["match_level"])})
        .groupby(["species", "Matching level"], sort=True)
        .size()
        .reset_index(name="Records")
        .rename(columns={"species": "Species"})
    )
    species_counts = truth["species"].value_counts()
    matching["% of species"] = matching.apply(
        lambda row: 100.0 * row["Records"] / species_counts[row["Species"]], axis=1
    )
    matching["% of benchmark"] = 100.0 * matching["Records"] / len(truth)

    source = (
        truth.groupby(["species", "source_database"], sort=True)
        .size()
        .reset_index(name="Records")
        .rename(columns={"species": "Species", "source_database": "Source database"})
    )
    source["% of species"] = source.apply(
        lambda row: 100.0 * row["Records"] / species_counts[row["Species"]], axis=1
    )
    source["% of benchmark"] = 100.0 * source["Records"] / len(truth)

    condition_rows = []
    for species, part in truth.groupby("species", sort=True):
        ph_n = int(part["pH"].notna().sum())
        temp_n = int(part["temperature_c"].notna().sum())
        condition_rows.append(
            {
                "Species": species,
                "Records": len(part),
                "pH available (n)": ph_n,
                "Temperature available (n)": temp_n,
                "pH available (%)": 100.0 * ph_n / len(part),
                "Temperature available (%)": 100.0 * temp_n / len(part),
            }
        )
    return matching, source, pd.DataFrame(condition_rows)


def subset_table(methods: dict[str, pd.DataFrame], identifiers: set[str], prefix: str) -> pd.DataFrame:
    rows = []
    for method in METHOD_ORDER:
        result = metrics(frame_by_ids(methods[method], identifiers))
        rows.append(
            {
                "subset": prefix,
                "method": method,
                "subset_total_rows": len(identifiers),
                "n": result["n"],
                "coverage_within_subset_percent": 100.0 * result["n"] / len(identifiers),
                **{key: value for key, value in result.items() if key != "n"},
            }
        )
    return pd.DataFrame(rows)


def build_reaction_subset(methods: dict[str, pd.DataFrame]) -> tuple[pd.DataFrame, set[str]]:
    identifiers = set.intersection(
        *(set(methods[method]["entry_id"].astype(str)) for method in REACTION_METHODS)
    )
    rows = []
    for method in METHOD_ORDER:
        result = metrics(frame_by_ids(methods[method], identifiers))
        rows.append(
            {
                "method": method,
                "common_subset_total_rows": len(identifiers),
                "n_common_subset": result["n"],
                "mae_log10_common_subset": result["mae_log10"],
                "bias_log10_common_subset": result["bias_log10"],
                "within_0.3_fraction_common_subset": result["within_0.3_log10_fraction"],
                "within_1.0_fraction_common_subset": result["within_1.0_log10_fraction"],
            }
        )
    return pd.DataFrame(rows), identifiers


def bh_adjust(p_values: list[float]) -> list[float]:
    n = len(p_values)
    order = np.argsort(p_values)
    adjusted = np.ones(n)
    running = 1.0
    for rank in range(n, 0, -1):
        index = int(order[rank - 1])
        running = min(running, p_values[index] * n / rank)
        adjusted[index] = min(1.0, running)
    return adjusted.tolist()


def build_wilcoxon(methods: dict[str, pd.DataFrame]) -> pd.DataFrame:
    rows = []
    p_values = []
    indexed = {method: frame.set_index("entry_id") for method, frame in methods.items()}
    for method_a, method_b in combinations(METHOD_ORDER, 2):
        left = indexed[method_a]
        right = indexed[method_b]
        common = left.index.intersection(right.index)
        error_a = left.loc[common, "abs_error_log10"].to_numpy(float)
        error_b = right.loc[common, "abs_error_log10"].to_numpy(float)
        result = wilcoxon(error_a, error_b, alternative="two-sided")
        p_value = float(result.pvalue)
        p_values.append(p_value)
        rows.append(
            {
                "method_a": method_a,
                "method_b": method_b,
                "n_common": len(common),
                "mae_a": float(error_a.mean()),
                "mae_b": float(error_b.mean()),
                "wilcoxon_statistic": float(result.statistic),
                "p_value_raw": p_value,
            }
        )
    adjusted = bh_adjust(p_values)
    for row, q_value in zip(rows, adjusted):
        row["p_value_bh"] = q_value
        row["significant_bh_fdr_0.05"] = bool(q_value < 0.05)
        row["better_method"] = row["method_a"] if row["mae_a"] < row["mae_b"] else row["method_b"]
    return pd.DataFrame(rows)


def build_stratification(
    methods: dict[str, pd.DataFrame], context: pd.DataFrame
) -> tuple[pd.DataFrame, pd.DataFrame]:
    selected_methods = ["KcatNet", "CataPro", "TurNuP", "PMAK", "KinForm-L", "CatPred"]
    context = context.copy()
    context["substrate_role_group_substrate_supported"] = context[
        "substrate_role_group"
    ].where(context["experimental_substrate_support"].eq("substrate_supported"), "outside_scope")
    feature_groups = {
        "species": ["ecoli", "yeast"],
        "source_database": ["BRENDA", "SABIO-RK", "BRENDA;SABIO-RK"],
        "experimental_substrate_support": ["substrate_supported", "participant_ambiguous"],
        "substrate_role_group": [
            "other_reactant",
            "currency_or_cofactor",
            "carrier_linked_variable",
        ],
        "substrate_role_group_substrate_supported": [
            "other_reactant",
            "currency_or_cofactor",
            "carrier_linked_variable",
        ],
    }
    rows = []
    for method in selected_methods:
        frame = methods[method][["entry_id", "abs_error_log10"]].merge(
            context[["entry_id", *feature_groups]], on="entry_id", validate="one_to_one"
        )
        for feature, groups in feature_groups.items():
            for group in groups:
                mask = frame[feature].astype(str).eq(str(group))
                part = frame[mask]
                rows.append(
                    {
                        "method": method,
                        "feature": feature,
                        "group": group,
                        "n": len(part),
                        "outlier_rows_abs_error_gt_1": int((part["abs_error_log10"] > 1.0).sum()),
                        "outlier_fraction": float((part["abs_error_log10"] > 1.0).mean()),
                        "mean_abs_error_log10": float(part["abs_error_log10"].mean()),
                    }
                )

    large_rows = []
    context_columns = [
        "entry_id",
        "species",
        "source_database",
        "experimental_substrate_support",
        "substrate_role_group",
        "currency_or_cofactor_like_by_name",
        "kegg_like_primary_group_short",
        "enzyme_complex_type",
        "n_measurements",
    ]
    for method in selected_methods:
        top = methods[method].nlargest(5, "abs_error_log10")[
            ["entry_id", "abs_error_log10"]
        ].merge(context[context_columns], on="entry_id", validate="one_to_one")
        top.insert(0, "method", method)
        large_rows.append(top)
    return pd.DataFrame(rows), pd.concat(large_rows, ignore_index=True)


def rank_stability(
    methods: dict[str, pd.DataFrame],
    method_names: list[str],
    comparison_set: str,
    replicates: int,
    seed: int,
) -> pd.DataFrame:
    common = set.intersection(*(set(methods[name]["entry_id"]) for name in method_names))
    identifiers = sorted(common)
    errors = np.vstack(
        [
            methods[name].set_index("entry_id").loc[identifiers, "abs_error_log10"].to_numpy(float)
            for name in method_names
        ]
    )
    rng = np.random.default_rng(seed)
    ranks = np.empty((replicates, len(method_names)))
    maes = np.empty((replicates, len(method_names)))
    for index in range(replicates):
        selected = rng.integers(0, len(identifiers), size=len(identifiers))
        replicate_mae = errors[:, selected].mean(axis=1)
        maes[index] = replicate_mae
        ranks[index] = rankdata(replicate_mae, method="average")
    rows = []
    for method_index, method in enumerate(method_names):
        rows.append(
            {
                "comparison_set": comparison_set,
                "common_n": len(identifiers),
                "method": method,
                "rank_1_bootstrap_frequency": float((ranks[:, method_index] == 1).mean()),
                "median_rank": float(np.median(ranks[:, method_index])),
                "rank_ci_low_95": float(np.percentile(ranks[:, method_index], 2.5)),
                "rank_ci_high_95": float(np.percentile(ranks[:, method_index], 97.5)),
                "bootstrap_mae_median": float(np.median(maes[:, method_index])),
                "bootstrap_replicates": replicates,
                "seed": seed,
            }
        )
    return pd.DataFrame(rows)


def paired_bootstrap(
    methods: dict[str, pd.DataFrame], identifiers: set[str], replicates: int, seed: int
) -> pd.DataFrame:
    ids = sorted(identifiers)
    errors = {
        method: methods[method].set_index("entry_id").loc[ids, "abs_error_log10"].to_numpy(float)
        for method in REACTION_METHODS
    }
    rng = np.random.default_rng(seed)
    selected = rng.integers(0, len(ids), size=(replicates, len(ids)))
    rows = []
    for method_a, method_b in combinations(REACTION_METHODS, 2):
        difference = errors[method_a].mean() - errors[method_b].mean()
        boot_difference = (
            errors[method_a][selected].mean(axis=1) - errors[method_b][selected].mean(axis=1)
        )
        low, high = np.percentile(boot_difference, [2.5, 97.5])
        rows.append(
            {
                "method_a": method_a,
                "method_b": method_b,
                "common_n": len(ids),
                "mae_difference_a_minus_b": float(difference),
                "bootstrap_ci_low_95": float(low),
                "bootstrap_ci_high_95": float(high),
                "bootstrap_frequency_method_a_lower_mae": float((boot_difference < 0).mean()),
                "bootstrap_replicates": replicates,
                "seed": seed,
            }
        )
    return pd.DataFrame(rows)


def build_overlap(methods: dict[str, pd.DataFrame]) -> pd.DataFrame:
    rows = []
    pretkcat = methods["PreTKcat"]
    pre_mask = bool_series(pretkcat["pretkcat_source_train_exact_pair_overlap"])
    rows.append(
        {
            "Method": "PreTKcat",
            "Overlap level": "Standardized sequence-parent pair in original public corpus (excluded before fitting)",
            "Overlap records": int(pre_mask.sum()),
            "Evaluated records": len(pretkcat),
            "MAE overlap": float(pretkcat.loc[pre_mask, "abs_error_log10"].mean()),
            "MAE non-overlap": float(pretkcat.loc[~pre_mask, "abs_error_log10"].mean()),
            "Within 10-fold overlap": float((pretkcat.loc[pre_mask, "abs_error_log10"] <= 1).mean()),
            "Within 10-fold non-overlap": float((pretkcat.loc[~pre_mask, "abs_error_log10"] <= 1).mean()),
            "Interpretation": (
                "All source rows matching the model sequence plus uncharged largest-fragment "
                "connectivity identity were removed before fitting; this reports source-corpus "
                "proximity, not residual training leakage."
            ),
        }
    )

    dekp = methods["DEKP-public-retrained"]
    dekp_mask = bool_series(dekp["dekp_train_exact_pair_overlap"])
    rows.append(
        {
            "Method": "DEKP-public-retrained",
            "Overlap level": "Standardized sequence-parent pair in original public corpus (excluded before fitting)",
            "Overlap records": int(dekp_mask.sum()),
            "Evaluated records": len(dekp),
            "MAE overlap": float(dekp.loc[dekp_mask, "abs_error_log10"].mean()),
            "MAE non-overlap": float(dekp.loc[~dekp_mask, "abs_error_log10"].mean()),
            "Within 10-fold overlap": float((dekp.loc[dekp_mask, "abs_error_log10"] <= 1).mean()),
            "Within 10-fold non-overlap": float((dekp.loc[~dekp_mask, "abs_error_log10"] <= 1).mean()),
            "Interpretation": (
                f"{int(pd.read_csv(DEKP_RUN_REPORT).iloc[0]['public_exact_pair_overlap_rows']):,} "
                "source rows matching full sequence plus uncharged largest-fragment connectivity "
                "identity were excluded before fitting."
            ),
        }
    )

    catpred_ids = set(methods["CatPred"]["entry_id"])
    overlap = pd.read_csv(CATPRED_OVERLAP)
    overlap = overlap[overlap["entry_id"].isin(catpred_ids)]
    levels = [
        ("Exact UniProt + reactant SMILES", "exact_uniprot_reactant_smiles_overlap"),
        ("Same UniProt + substrate component", "same_uniprot_substrate_component_overlap"),
        ("UniProt overlap", "uniprot_overlap"),
        ("EC overlap", "ec_overlap"),
    ]
    interpretations = {
        "Exact UniProt + reactant SMILES": "Few exact raw-string matches against the CatPred reference corpus.",
        "Same UniProt + substrate component": "Broader enzyme-substrate component proximity.",
        "UniProt overlap": "Protein identity overlap is common.",
        "EC overlap": "Functional-class overlap covers most evaluated rows.",
    }
    for level, column in levels:
        rows.append(
            {
                "Method": "CatPred",
                "Overlap level": level,
                "Overlap records": int(bool_series(overlap[column]).sum()),
                "Evaluated records": len(overlap),
                "MAE overlap": np.nan,
                "MAE non-overlap": np.nan,
                "Within 10-fold overlap": np.nan,
                "Within 10-fold non-overlap": np.nan,
                "Interpretation": interpretations[level],
            }
        )
    return pd.DataFrame(rows)


def software_versions(release: dict[str, object]) -> pd.DataFrame:
    commits = {
        "CatPred": "8e72d324e9e6f7a9a24c3f8a720884c7c1740a9b",
        "CataPro": "cc89b2c81768665cf6fd76dfda607ce88691f601c",
        "DEKP-public-retrained": "d2b8c1372b5c1855fd2de9aaadde19cf8cc7fa8d",
        "DLKcat": "7c15d0d4a7ac029f9d75564d9f2a93874aeaaec7",
        "GO-HKP": "5d086a4ded543295250eb7db2a1ea4b1336e7f48",
        "KcatNet": "7d370f69f9d1bbed517655d23d4d80bd76594321",
        "KinForm-L": "f7a70eb1cd6723ba3a8d606432e522ea2b0fa9fd",
        "PMAK": "1b1bea4580ef7bb908f893d3b13213a1486bbb98",
        "PreTKcat": "b7bc0562a9b8555a201c5f6c72fc2a660dcdb76d",
        "SELFprot": "880c2e8fd685ed0e8d574382439f7d7ca75cc9d0",
        "UniKP": "5cee5c4a64ba2daf59c63a5b5cbaa0cadf97ef26",
        "TurNuP": "local source archive; no Git commit recorded",
    }
    runtime = {
        "CatPred": "Python 3.11.13; PyTorch 2.6.0+cu124; scikit-learn 1.7.1; transformers 4.51.3",
        "KinForm-L": "Python 3.12.11; PyTorch 2.11.0+cu130; scikit-learn 1.7.2; joblib 1.4.2",
        "UniKP": "features: Python 3.10.13; prediction: Python 3.6.12, scikit-learn 0.23.1",
    }
    shared = (
        "Python 3.10.13; PyTorch 2.1.2.post301; scikit-learn 1.3.2; "
        "transformers 4.35.2; RDKit 2026.03.3"
    )
    rows = []
    for method in METHOD_ORDER:
        rows.append(
            {
                "Method": method,
                "Upstream revision": commits[method],
                "Runtime environment": runtime.get(method, shared),
                "Benchmark version": release["benchmark_version"],
                "Evaluation date": release["release_date"],
            }
        )
    return pd.DataFrame(rows)


def release_info(release: dict[str, object]) -> pd.DataFrame:
    return pd.DataFrame(
        [
            ("Benchmark", release["benchmark_name"]),
            ("Benchmark version", release["benchmark_version"]),
            ("Table schema version", release["schema_version"]),
            ("Release date", release["release_date"]),
            ("Data freeze date", release["data_freeze_date"]),
            ("Artifact revision", release["artifact_revision"]),
            ("Artifact revision date", release["artifact_revision_date"]),
            ("Canonical rows", release["canonical_rows"]),
            ("E. coli rows", release["species_rows"]["ecoli"]),
            ("Yeast rows", release["species_rows"]["yeast"]),
            ("Canonical benchmark SHA256", release["canonical_benchmark_sha256"]),
            (
                "E. coli source model",
                f"{release['source_models']['ecoli']['model_id']} version {release['source_models']['ecoli']['version']}",
            ),
            ("E. coli model SHA256", release["source_models"]["ecoli"]["sha256"]),
            (
                "Yeast source model",
                f"{release['source_models']['yeast']['model_id']} version {release['source_models']['yeast']['version']}",
            ),
            ("Yeast model SHA256", release["source_models"]["yeast"]["sha256"]),
            ("BRENDA release", release["data_sources"]["BRENDA"]["release"]),
            ("SABIO-RK data freeze", release["data_sources"]["SABIO-RK"]["data_freeze_date"]),
            ("Target space", release["statistics"]["target_space"]),
            ("Bootstrap replicates", release["statistics"]["bootstrap_replicates"]),
            ("Conditional row-bootstrap seed", release["statistics"]["row_bootstrap_seed"]),
            ("Cluster-bootstrap seed", release["statistics"]["cluster_bootstrap_seed"]),
            ("Primary uncertainty", release["statistics"]["primary_uncertainty"]),
            ("BRENDA archive SHA256", release["data_sources"]["BRENDA"]["archive_sha256"]),
            ("SABIO-RK parsed-table SHA256", release["data_sources"]["SABIO-RK"]["parsed_table_sha256"]),
            ("CKB snapshot commit", release["data_sources"]["CKB"]["commit"]),
            ("CKB database SHA256", release["data_sources"]["CKB"]["database_sha256"]),
            (
                "PreTKcat training policy",
                "exact sequence-parent identity excluded for the primary reconstruction; raw and near-excluded variants reported in S24",
            ),
            ("Quinate structure source", "PubChem CID 6508; neutral C7H12O6 representation"),
        ],
        columns=["Field", "Value"],
    )


def write_frame(workbook: Workbook, name: str, frame: pd.DataFrame, index: int | None = None) -> None:
    if name in workbook.sheetnames:
        workbook.remove(workbook[name])
    worksheet = workbook.create_sheet(name, index if index is not None else len(workbook.sheetnames))
    worksheet.append(list(frame.columns))
    for row in frame.itertuples(index=False, name=None):
        worksheet.append([None if pd.isna(value) else value for value in row])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    worksheet.row_dimensions[1].height = 34
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for column_index, column in enumerate(frame.columns, start=1):
        values = [str(column)] + [str(value) for value in frame[column].dropna().head(1000)]
        width = min(55, max(11, max(len(value) for value in values) + 2))
        worksheet.column_dimensions[get_column_letter(column_index)].width = width
        column_name = str(column).lower()
        for cell in worksheet.iter_cols(
            min_col=column_index,
            max_col=column_index,
            min_row=2,
            max_row=worksheet.max_row,
        ):
            for item in cell:
                if isinstance(item.value, (float, np.floating)):
                    if "p_value" in column_name or column_name in {"p", "q"}:
                        item.number_format = "0.000000"
                    elif "percent" in column_name or column_name.endswith("_pct"):
                        item.number_format = "0.0"
                    else:
                        item.number_format = "0.000"
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=False)


def style_supplementary_audits(workbook: Workbook) -> None:
    s20 = workbook["S20_Training_overlap"]
    s20.freeze_panes = "D2"
    s20.sheet_view.zoomScale = 70
    for column_index in range(1, s20.max_column + 1):
        letter = get_column_letter(column_index)
        s20.column_dimensions[letter].width = min(s20.column_dimensions[letter].width or 16, 18)

    specifications = {
        "S17_Sensitivity_subsets": (
            {"analysis_scope": 38, "statistical_unit": 34, "aggregation_rule": 58},
            48,
        ),
        "S20_Training_overlap": (
            {
                "method": 24,
                "training_corpus_status": 26,
                "training_corpus_source": 46,
                "chemical_identity": 42,
                "publication_overlap_status": 26,
                "interpretation": 58,
            },
            64,
        ),
        "S21_Measurement_dispersion": (
            {
                "analysis": 42,
                "denominator": 34,
                "dispersion_formula": 58,
                "interpretation": 58,
            },
            64,
        ),
        "S22_Mutation_status": (
            {
                "audit_status": 30,
                "counting_stage": 42,
                "benchmark_policy": 58,
            },
            58,
        ),
        "S23_Substrate_direction": (
            {"audit_item": 46, "denominator": 22, "interpretation": 60},
            54,
        ),
        "S24_PreTKcat_variants": (
            {
                "variant": 22,
                "manuscript_role": 28,
                "training_overlap_policy": 42,
                "pair_identity_definition": 48,
                "near_neighbor_definition": 58,
                "audit_file": 52,
                "metrics_file": 52,
            },
            64,
        ),
    }
    for sheet_name, (column_widths, row_height) in specifications.items():
        worksheet = workbook[sheet_name]
        header_columns = {cell.value: cell.column for cell in worksheet[1]}
        for header, width in column_widths.items():
            column_index = header_columns[header]
            letter = get_column_letter(column_index)
            worksheet.column_dimensions[letter].width = width
            for row_index in range(2, worksheet.max_row + 1):
                worksheet.cell(row_index, column_index).alignment = Alignment(
                    vertical="top", wrap_text=True
                )
        for row_index in range(2, worksheet.max_row + 1):
            worksheet.row_dimensions[row_index].height = row_height


def style_table0(workbook: Workbook) -> None:
    worksheet = workbook["Table0"]
    worksheet.freeze_panes = "F2"
    worksheet.sheet_view.zoomScale = 70
    headers = {cell.value: cell.column for cell in worksheet[1]}
    widths = {
        "method": 24,
        "inference_regime": 38,
        "inference_time_modality": 38,
        "prediction_status": 32,
        "entry_id": 36,
        "benchmark_sequence": 52,
        "method_sequence_input": 52,
        "method_sequence_status": 32,
        "sequence_input_policy": 58,
        "reaction_id": 22,
        "reaction_name": 42,
        "reaction_equation_ids_model_forward": 62,
        "reaction_equation_names_model_forward": 70,
        "reaction_input_role": 58,
        "method_reaction_input": 62,
        "evaluated_metabolite_id": 28,
        "evaluated_metabolite_name": 38,
        "evaluated_metabolite_smiles": 55,
        "candidate_selection_policy": 52,
        "substrate_parent_inchikey_connectivity": 34,
        "substrate_role_evidence": 58,
        "substrate_role_confidence": 34,
        "experimental_substrate_support": 30,
        "experimental_source_record_ids": 58,
        "experimental_source_database": 28,
        "experimental_match_level": 38,
    }
    for header, width in widths.items():
        worksheet.column_dimensions[get_column_letter(headers[header])].width = width
    for header in [
        "experimental_kcat_s^-1",
        "predicted_kcat_s^-1",
    ]:
        column = headers[header]
        for row in range(2, worksheet.max_row + 1):
            worksheet.cell(row, column).number_format = "0.000000E+00"


def style_table0_wide(workbook: Workbook) -> None:
    worksheet = workbook["Table0_wide"]
    worksheet.freeze_panes = "I2"
    worksheet.column_dimensions["A"].width = 34
    worksheet.column_dimensions["G"].width = 28
    worksheet.column_dimensions["H"].width = 50
    for column in range(9, worksheet.max_column + 1):
        worksheet.column_dimensions[get_column_letter(column)].width = 24
        for row in range(2, worksheet.max_row + 1):
            worksheet.cell(row, column).number_format = "0.000000E+00"


def validate_workbook(path: Path, expected_sheets: list[str]) -> None:
    """Reject incomplete or malformed OOXML before replacing the public artifact."""
    with ZipFile(path) as archive:
        bad_member = archive.testzip()
        if bad_member is not None:
            raise ValueError(f"Corrupt XLSX member: {bad_member}")
        for member in archive.namelist():
            if member.endswith(".xml"):
                ElementTree.fromstring(archive.read(member))

    reopened = load_workbook(path, read_only=False, data_only=False)
    try:
        if reopened.sheetnames != expected_sheets:
            raise ValueError(f"Unexpected worksheet order: {reopened.sheetnames}")
        expected_table0_rows = 1 + len(METHOD_ORDER) * BENCHMARK_N
        if reopened["Table0"].max_row != expected_table0_rows or reopened["Table0"].max_column != 41:
            raise ValueError("Table0 dimensions changed during workbook serialization")
        if reopened["Table0_wide"].max_row != BENCHMARK_N + 1 or reopened["Table0_wide"].max_column != 22:
            raise ValueError("Table0_wide dimensions changed during workbook serialization")
        if any(worksheet.tables for worksheet in reopened.worksheets):
            raise ValueError("Compatibility workbook unexpectedly contains table objects")
        if any(
            cell.comment is not None
            for worksheet in reopened.worksheets
            for row in worksheet.iter_rows()
            for cell in row
        ):
            raise ValueError("Compatibility workbook unexpectedly contains VML comments")
    finally:
        reopened.close()


def build_index(frames: dict[str, pd.DataFrame]) -> pd.DataFrame:
    titles = {
        "Release_info": "Benchmark version, dates, checksum, and statistical settings.",
        "Table0": (
            f"Method-level long-format input and prediction audit for all 12 methods x {BENCHMARK_N} benchmark "
            "records, including sequence policy, model-forward reaction equation, evaluated metabolite, "
            "role evidence, prediction status, experimental kcat, and predicted kcat."
        ),
        "Table0_wide": f"Wide {BENCHMARK_N}-row prediction matrix with one prediction column per method.",
        "Record_audit": (
            "Per-record provenance, label-dependence, measurement, mapping, model-direction, "
            "substrate-role, and method-specific training-proximity fields needed to reproduce S16-S24."
        ),
        "Table1": "Inference regimes used for coverage-aware comparison.",
        "Table2": "Performance within explicitly defined evaluation scopes, using pair-cluster confidence intervals.",
        "S1_Full_metrics": "Complete overall metrics for each achieved evaluation set.",
        "S2_Method_details": "Method inputs, implementation status, preprocessing, and caveats.",
        "S3_Matching_levels": "Experimental matching levels by species.",
        "S4_Source_support": "BRENDA and SABIO-RK support by species.",
        "S5_Condition_metadata": "pH and temperature availability.",
        "S6_Reaction_subset": "Metrics on the strict common reaction-aware subset.",
        "S7_Available_case": (
            "Available-case metrics within CatPred- and KinForm-L-accessible scopes; cell-specific n is required."
        ),
        "S8_Wilcoxon": "Pairwise Wilcoxon tests with global BH-adjusted q values.",
        "S9_Error_stratification": (
            "Error by species, source support, and registry-defined substrate-role group."
        ),
        "S10_Large_errors": "Five largest errors for representative methods.",
        "S11_Bootstrap_CI": "Conditional row-resampling confidence intervals for four performance metrics.",
        "S12_Rank_stability": "Paired row-resampling bootstrap rank-one frequencies and rank stability.",
        "S13_Paired_bootstrap": "Paired row-resampling MAE-difference bootstrap on the reaction-aware subset.",
        "S14_Overlap_audit": "Method-specific public-corpus proximity and exclusion audit.",
        "S15_Software_versions": "Upstream revisions, runtime versions, and evaluation date.",
        "S16_Label_audit": "Record matching strength, label-assignment multiplicity, and dependence audit.",
        "S17_Sensitivity_subsets": (
            "Sensitivity metrics, including strictly common reaction-set unique-pair and unique-label analyses; "
            "observed and predicted log10(kcat) are separately median-aggregated within each cluster and clusters are weighted equally."
        ),
        "S18_Cluster_bootstrap": (
            "Protein-, pair-, reaction-, reference-, and label-assignment-cluster bootstrap intervals, "
            "including the SABIO participant-ambiguity exclusion scope."
        ),
        "S19_Cluster_wilcoxon": "Cluster-aggregated paired comparisons with BH adjustment.",
        "S20_Training_overlap": (
            "Standardized exact, joint-near-neighbor, and no-joint-neighbor-under-thresholds "
            "training-corpus proximity and performance audit."
        ),
        "S21_Measurement_dispersion": (
            "Within-entry and between-source experimental kcat dispersion with explicit formulas."
        ),
        "S22_Mutation_status": (
            "Mutation and variant-status audit with before-filter, excluded, and retained stages."
        ),
        "S23_Substrate_direction": (
            "Substrate-role, model encoding, model reversibility, assay-direction limitation, and currency audit."
        ),
        "S24_PreTKcat_variants": (
            "Raw-public, exact-overlap-excluded primary, and joint-near-neighbor-excluded PreTKcat reconstructions."
        ),
    }
    rows = []
    for name in frames:
        label = name
        if name.startswith("S"):
            label = "Supplementary Table " + name.split("_", 1)[0]
        elif name == "Table0":
            label = "Supplementary Data Table 0"
        elif name == "Table0_wide":
            label = "Supplementary Data Table 0W"
        elif name == "Record_audit":
            label = "Supplementary Record Audit"
        rows.append({"Table": label, "Title": titles[name], "Worksheet": name})
    return pd.DataFrame(rows)


def main() -> None:
    args = parse_args()
    release = json.loads(RELEASE.read_text(encoding="utf-8"))
    truth = pd.read_csv(TRUTH)
    context = pd.read_csv(CONTEXT)
    methods = load_methods()
    method_inputs = load_method_inputs()
    reaction_provenance = build_reaction_provenance()
    audit_tables = load_audit_tables()
    record_audit_path = TABLE_EXPORT_DIR / "Record_audit.csv"
    if not record_audit_path.exists():
        raise FileNotFoundError(
            f"Missing {record_audit_path}; run paper/build_submission_audits.py first"
        )
    record_audit = pd.read_csv(record_audit_path, low_memory=False)

    bootstrap = pd.concat(
        [
            bootstrap_method(methods[method], method, args.bootstrap_replicates, args.seed)
            for method in METHOD_ORDER
        ],
        ignore_index=True,
    )
    full = build_full_metrics(methods)
    table2 = build_table2(methods, audit_tables["S18_Cluster_bootstrap"])
    matching, source, condition = build_provenance(truth)
    reaction_subset, reaction_ids = build_reaction_subset(methods)
    catpred_scope_n = len(methods["CatPred"])
    kinform_scope_n = len(methods["KinForm-L"])
    reaction_scope_n = len(reaction_ids)
    matched = pd.concat(
        [
            subset_table(
                methods,
                set(methods["CatPred"]["entry_id"]),
                f"catpred_accessible_scope_{catpred_scope_n}_available_case",
            ),
            subset_table(
                methods,
                set(methods["KinForm-L"]["entry_id"]),
                f"kinform_accessible_scope_{kinform_scope_n}_available_case",
            ),
        ],
        ignore_index=True,
    )
    stratification, large_errors = build_stratification(methods, context)
    ranks = pd.concat(
        [
            rank_stability(
                methods,
                NEAR_FULL,
                f"broad_common_{BENCHMARK_N}",
                args.bootstrap_replicates,
                args.seed,
            ),
            rank_stability(
                methods,
                REACTION_METHODS,
                f"reaction_aware_common_{reaction_scope_n}",
                args.bootstrap_replicates,
                args.seed,
            ),
        ],
        ignore_index=True,
    )

    frames = {
        "Release_info": release_info(release),
        "Table0": build_table0(
            truth, context, methods, method_inputs, reaction_provenance
        ),
        "Table0_wide": build_table0_wide(truth, methods),
        "Record_audit": record_audit,
        "Table1": build_table1(methods),
        "Table2": table2,
        "S1_Full_metrics": full,
        "S2_Method_details": build_method_details(methods),
        "S3_Matching_levels": matching,
        "S4_Source_support": source,
        "S5_Condition_metadata": condition,
        "S6_Reaction_subset": reaction_subset,
        "S7_Available_case": matched,
        "S8_Wilcoxon": build_wilcoxon(methods),
        "S9_Error_stratification": stratification,
        "S10_Large_errors": large_errors,
        "S11_Bootstrap_CI": bootstrap,
        "S12_Rank_stability": ranks,
        "S13_Paired_bootstrap": paired_bootstrap(
            methods, reaction_ids, args.bootstrap_replicates, args.seed
        ),
        "S14_Overlap_audit": build_overlap(methods),
        "S15_Software_versions": software_versions(release),
        **audit_tables,
    }

    workbook = Workbook()
    workbook.remove(workbook.active)
    write_frame(workbook, "Index", build_index(frames))
    for name, frame in frames.items():
        write_frame(workbook, name, frame)
    style_table0(workbook)
    style_table0_wide(workbook)
    style_supplementary_audits(workbook)
    workbook.properties.title = "kcat benchmark manuscript tables"
    workbook.properties.subject = (
        f"Benchmark v{release['benchmark_version']}; data freeze {release['data_freeze_date']}"
    )
    workbook.properties.creator = "kcat benchmark analysis pipeline"
    workbook.properties.keywords = (
        f"kcat, benchmark, version {release['benchmark_version']}, {release['data_freeze_date']}"
    )

    args.output.parent.mkdir(parents=True, exist_ok=True)
    temporary_output = args.output.with_name(f".{args.output.name}.tmp.xlsx")
    workbook.save(temporary_output)
    workbook.close()
    validate_workbook(temporary_output, ["Index", *frames])
    temporary_output.replace(args.output)

    TABLE_EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    for name, frame in frames.items():
        frame.to_csv(TABLE_EXPORT_DIR / f"{name}.csv", index=False)
    build_index(frames).to_csv(TABLE_EXPORT_DIR / "Index.csv", index=False)

    snapshot = {
        "benchmark_version": release["benchmark_version"],
        "data_freeze_date": release["data_freeze_date"],
        "canonical_rows": len(truth),
        "method_rows": {method: len(frame) for method, frame in methods.items()},
        "reaction_subset_rows": len(reaction_ids),
        "catpred_subset_rows": len(methods["CatPred"]),
        "kinform_subset_rows": len(methods["KinForm-L"]),
        "bootstrap_replicates": args.bootstrap_replicates,
        "bootstrap_seed": args.seed,
        "overall_metrics": {
            row["Method"]: {
                "n": int(row["n"]),
                "mae_log10": float(row["MAE log10"]),
                "rmse_log10": float(row["RMSE log10"]),
                "spearman_log10": float(row["Spearman"]),
            }
            for _, row in full.iterrows()
        },
    }
    SNAPSHOT.write_text(json.dumps(snapshot, indent=2) + "\n", encoding="ascii")
    print(f"Wrote workbook: {args.output}")
    print(f"Wrote CSV tables: {TABLE_EXPORT_DIR}")
    print(f"Wrote statistics snapshot: {SNAPSHOT}")


if __name__ == "__main__":
    main()
