#!/usr/bin/env python3
"""Add a per-record kcat prediction matrix as worksheet Table0."""

from __future__ import annotations

import argparse
import math
from copy import copy
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


BASE = Path(__file__).resolve().parent.parent
DEFAULT_INPUT = BASE / "paper" / "kcat_benchmark_reorganized_tables.xlsx"
DEFAULT_OUTPUT = BASE / "paper" / "kcat_benchmark_reorganized_tables_reviewed.xlsx"
TRUTH_PATH = BASE / "data" / "final" / "benchmark_ready_catpred.csv"
SUMMARY_PATH = BASE / "reports" / "tables" / "method_eval_summary.csv"

METHOD_FILES = {
    "KcatNet": BASE / "data" / "final" / "kcatnet" / "kcatnet_kcat_predictions_evaluated.csv",
    "CataPro": BASE / "data" / "final" / "catapro" / "catapro_kcat_predictions_evaluated.csv",
    "PreTKcat": BASE / "data" / "final" / "pretkcat" / "pretkcat_kcat_predictions_evaluated.csv",
    "UniKP": BASE / "data" / "final" / "unikp" / "unikp_kcat_predictions_evaluated.csv",
    "SELFprot": BASE / "data" / "final" / "selfprot" / "selfprot_kcat_predictions_evaluated.csv",
    "DLKcat": BASE / "data" / "final" / "dlkcat" / "dlkcat_kcat_predictions_evaluated.csv",
    "TurNuP": BASE / "data" / "final" / "turnup" / "turnup_kcat_predictions_evaluated.csv",
    "PMAK": BASE / "data" / "final" / "pmak" / "pmak_kcat_predictions_evaluated.csv",
    "KinForm": BASE / "data" / "final" / "kinform" / "kinform_kcat_predictions_evaluated.csv",
    "CatPred": BASE / "data" / "final" / "catpred" / "catpred_kcat_predictions_evaluated.csv",
    "DEKP-public-retrained": BASE
    / "data"
    / "final"
    / "dekp"
    / "dekp_public_retrained_kcat_predictions_evaluated.csv",
    "GO-HKP": BASE / "data" / "final" / "go_hkp" / "go_hkp_kcat_predictions_evaluated.csv",
}

DISPLAY_NAMES = {
    "KinForm": "KinForm-L_predicted_kcat_s^-1",
    "DEKP-public-retrained": "DEKP_public_retrained_predicted_kcat_s^-1",
}

SUMMARY_NAMES = {
    "DLKcat": "DLKcat-official",
    "UniKP": "UniKP-official",
    "TurNuP": "TurNuP-official",
}

METHOD_COMMENTS = {
    "KcatNet": "Official public checkpoint; 26 sequences were truncated to the model limit. One invalid Quinate SMILES was not scored.",
    "CataPro": "Mean prediction from 10 released fold-specific models; all benchmark rows were treated as wild type. One invalid Quinate SMILES was not scored.",
    "PreTKcat": "Public-data retraining because no fitted kcat regressor was released. ExtraTrees used 16,249 public rows; 148 benchmark temperatures were imputed to 30 C, 26 sequences were truncated, and 26 benchmark pairs overlapped the public fitting corpus.",
    "UniKP": "Official public regressor; sequences longer than 1,000 residues were represented by their first and last 500 residues. One invalid Quinate SMILES was not scored.",
    "SELFprot": "Released fold-1 checkpoint only; one invalid Quinate SMILES was not scored.",
    "DLKcat": "Official public code and checkpoint; one invalid Quinate SMILES was not scored.",
    "TurNuP": "Official public workflow requiring complete reactant and product representations; 198 rows without complete reaction SMILES were outside its input domain.",
    "PMAK": "Mean prediction from five released reaction-cold fold checkpoints; 198 rows without complete reaction SMILES were outside its input domain.",
    "KinForm": "KinForm-L evaluation using released precomputed feature assets; 415 rows lacked the required lookup/assets or had an invalid SMILES.",
    "CatPred": "Production kcat ensemble; 64 monatomic-proton rows and one invalid Quinate SMILES did not yield predictions.",
    "DEKP-public-retrained": "Public-data retraining because final kcat weights were not released. Sixteen source rows matching benchmark pairs were excluded before fitting; one invalid Quinate SMILES was not scored.",
    "GO-HKP": "Functional-assignment baseline. E. coli used GO-HKP DeepGO-SE reaction assignments; yeast used UniProt GO terms and an organism-filtered GO-kcat table.",
}

IDENTIFIER_COLUMNS = [
    "entry_id",
    "species",
    "reaction_id",
    "gene_id",
    "uniprot_id",
    "ec_number",
    "substrate_name",
    "SMILES",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    return parser.parse_args()


def load_prediction_matrix() -> tuple[pd.DataFrame, dict[str, int]]:
    truth = pd.read_csv(TRUTH_PATH)
    if len(truth) != 978 or truth["entry_id"].duplicated().any():
        raise ValueError("The canonical truth table must contain 978 unique entry_id values.")

    table = truth[IDENTIFIER_COLUMNS + ["true_kcat", "true_kcat_log10"]].copy()
    table = table.rename(
        columns={
            "SMILES": "substrate_smiles",
            "true_kcat": "experimental_kcat_s^-1",
            "true_kcat_log10": "experimental_log10_kcat",
        }
    )
    coverage: dict[str, int] = {}

    for method, path in METHOD_FILES.items():
        rows = pd.read_csv(path, usecols=["entry_id", "prediction_kcat", "prediction_log10"])
        if rows["entry_id"].duplicated().any():
            raise ValueError(f"{method} contains duplicate entry_id values: {path}")
        linear_from_log = 10.0 ** pd.to_numeric(rows["prediction_log10"], errors="coerce")
        linear = pd.to_numeric(rows["prediction_kcat"], errors="coerce")
        if not ((linear - linear_from_log).abs() <= 1e-8 * linear_from_log.abs().clip(lower=1.0)).all():
            raise ValueError(f"{method} has inconsistent linear and log10 prediction columns.")

        output_column = DISPLAY_NAMES.get(method, f"{method}_predicted_kcat_s^-1")
        values = rows[["entry_id", "prediction_kcat"]].rename(columns={"prediction_kcat": output_column})
        table = table.merge(values, on="entry_id", how="left", validate="one_to_one")
        coverage[method] = int(table[output_column].notna().sum())

    return table, coverage


def validate_against_summary(table: pd.DataFrame, coverage: dict[str, int]) -> None:
    summary = pd.read_csv(SUMMARY_PATH).set_index("method")
    truth_log = pd.to_numeric(table["experimental_log10_kcat"], errors="coerce")
    for method in METHOD_FILES:
        summary_name = SUMMARY_NAMES.get(method, method)
        expected = summary.loc[summary_name]
        column = DISPLAY_NAMES.get(method, f"{method}_predicted_kcat_s^-1")
        predicted = pd.to_numeric(table[column], errors="coerce")
        valid = truth_log.notna() & predicted.gt(0)
        mae = (predicted.loc[valid].map(math.log10) - truth_log.loc[valid]).abs().mean()
        if coverage[method] != int(expected["n"]):
            raise ValueError(
                f"{method} coverage mismatch: Table0={coverage[method]}, summary={int(expected['n'])}"
            )
        if not math.isclose(mae, float(expected["mae_log10"]), rel_tol=0, abs_tol=1e-10):
            raise ValueError(
                f"{method} MAE mismatch: Table0={mae:.12g}, summary={float(expected['mae_log10']):.12g}"
            )


def style_table0(ws, coverage: dict[str, int]) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    missing_fill = PatternFill("solid", fgColor="FFF2CC")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 42
    ws.freeze_panes = "I2"
    ws.auto_filter.ref = ws.dimensions

    widths = {
        "A": 33,
        "B": 13,
        "C": 20,
        "D": 16,
        "E": 14,
        "F": 14,
        "G": 30,
        "H": 48,
        "I": 21,
        "J": 22,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for col_idx in range(11, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 25

    for row in ws.iter_rows(min_row=2, min_col=9, max_col=ws.max_column):
        for cell in row:
            cell.number_format = "0.000000E+00" if cell.column != 10 else "0.000000"
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top")

    ws["H1"].comment = Comment(
        "Substrate representation from the canonical benchmark input. One Quinate row contains the malformed numeric string 192.167 rather than a parseable SMILES; it is highlighted below and was not scored by structure-dependent methods.",
        "kcat benchmark review",
    )
    for row in range(2, ws.max_row + 1):
        if str(ws.cell(row, 8).value) == "192.167":
            ws.cell(row, 8).fill = PatternFill("solid", fgColor="F4CCCC")
            ws.cell(row, 8).comment = Comment(
                "Malformed legacy compound mapping. Correct this structure and rerun affected methods before treating the benchmark as 978-row sequence+SMILES complete.",
                "kcat benchmark review",
            )

    for col_idx, method in enumerate(METHOD_FILES, start=11):
        header = ws.cell(1, col_idx)
        header.comment = Comment(
            f"Predicted kcat in s^-1. Blank cells were not scored by this method. "
            f"Available predictions: {coverage[method]}/978. {METHOD_COMMENTS[method]}",
            "kcat benchmark review",
        )
        letter = get_column_letter(col_idx)
        ws.conditional_formatting.add(
            f"{letter}2:{letter}{ws.max_row}",
            FormulaRule(formula=[f'ISBLANK({letter}2)'], fill=missing_fill),
        )

    table = Table(displayName="PerRecordKcatPredictions", ref=ws.dimensions)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def update_index(wb) -> None:
    if "Index" not in wb.sheetnames:
        return
    ws = wb["Index"]
    existing = [ws.cell(row, 3).value for row in range(2, ws.max_row + 1)]
    if "Table0" in existing:
        return
    ws.insert_rows(2)
    ws.cell(2, 1, "Table 0")
    ws.cell(2, 2, "Per-record experimental and predicted kcat values for all 12 methods.")
    ws.cell(2, 3, "Table0")
    for col in range(1, 4):
        source = ws.cell(3, col)
        target = ws.cell(2, col)
        if source.has_style:
            target._style = copy(source._style)
        target.number_format = source.number_format
        target.alignment = copy(source.alignment)

    titles = {
        "S2_Method_details": "Method inputs, implementation status, preprocessing, and principal caveats.",
        "S8_Wilcoxon": "Pairwise Wilcoxon tests with raw P values and Benjamini-Hochberg-adjusted q values.",
        "S14_Overlap_audit": "Method-specific public-corpus proximity and overlap audit.",
    }
    for row in range(2, ws.max_row + 1):
        sheet = ws.cell(row, 3).value
        if sheet in titles:
            ws.cell(row, 2, titles[sheet])


def benjamini_hochberg(p_values: list[float]) -> list[float]:
    """Return monotone Benjamini-Hochberg q values in input order."""
    n = len(p_values)
    order = sorted(range(n), key=lambda index: p_values[index])
    adjusted = [1.0] * n
    running = 1.0
    for rank_from_end, index in enumerate(reversed(order), start=1):
        rank = n - rank_from_end + 1
        running = min(running, p_values[index] * n / rank)
        adjusted[index] = min(1.0, running)
    return adjusted


def update_wilcoxon_sheet(ws) -> None:
    if ws.cell(1, 7).value == "p_value_raw":
        return
    p_values = [float(ws.cell(row, 7).value) for row in range(2, ws.max_row + 1)]
    q_values = benjamini_hochberg(p_values)
    ws.insert_cols(8, 1)
    ws.cell(1, 7, "p_value_raw")
    ws.cell(1, 8, "p_value_bh")
    ws.cell(1, 9, "significant_bh_fdr_0.05")
    ws.cell(1, 10, "better_method")
    for row, q_value in enumerate(q_values, start=2):
        ws.cell(row, 7, float(p_values[row - 2]))
        ws.cell(row, 8, q_value)
        ws.cell(row, 9, bool(q_value < 0.05))
        for col in (7, 8):
            ws.cell(row, col).number_format = "0.000E+00"
    for cell in ws[1]:
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.column_dimensions["G"].width = 16
    ws.column_dimensions["H"].width = 16
    ws.column_dimensions["I"].width = 24


def update_method_sheets(wb) -> None:
    table1 = wb["Table1"]
    table1.cell(2, 1, "Near-full sequence+substrate")
    table1.cell(2, 3, "Enzyme sequence and one substrate representation; PreTKcat also uses temperature")
    table1.cell(2, 6, "No product-side information; PreTKcat is public-data retrained and temperature-aware.")
    table1.cell(4, 2, "CatPred; KinForm-L")
    table1.cell(4, 3, "Sequence and substrate plus method-specific preprocessing or released feature assets")
    table1.cell(4, 5, "Interpret within each achieved applicability domain")
    table1.cell(4, 6, "CatPred excludes 64 proton rows plus one invalid SMILES; KinForm-L depends on released precomputed assets.")
    table1.cell(5, 2, "DEKP-public-retrained")
    table1.cell(5, 6, "Final kcat weights were not released; exact benchmark pairs were excluded before retraining.")

    table2 = wb["Table2"]
    for row in range(2, table2.max_row + 1):
        if table2.cell(row, 2).value == "Near-full sequence+SMILES":
            table2.cell(row, 2, "Near-full sequence+substrate")
        if table2.cell(row, 1).value == "DEKP":
            table2.cell(row, 1, "DEKP-public-retrained")

    full_metrics = wb["S1_Full_metrics"]
    for row in range(2, full_metrics.max_row + 1):
        if full_metrics.cell(row, 13).value == "Near-full sequence+SMILES":
            full_metrics.cell(row, 13, "Near-full sequence+substrate")
        if full_metrics.cell(row, 1).value == "DEKP":
            full_metrics.cell(row, 1, "DEKP-public-retrained")
        if full_metrics.cell(row, 1).value == "PreTKcat":
            full_metrics.cell(row, 14, "sequence + substrate SMILES + temperature")

    details = wb["S2_Method_details"]
    replacements = {
        "CataPro": [
            "Near-full sequence+substrate",
            "Enzyme sequence; substrate SMILES; variant type set to wild",
            "ProtT5/MolT5-derived features, molecular fingerprints, and neural prediction",
            "Mean of 10 released fold-specific models",
            "One invalid Quinate SMILES omitted; no product-side input.",
        ],
        "DLKcat": [
            "Near-full sequence+substrate",
            "Enzyme sequence; substrate name/SMILES",
            "Substrate molecular graph with protein sequence CNN/attention",
            "Official code and checkpoint",
            "One invalid Quinate SMILES omitted; no product-side input.",
        ],
        "KcatNet": [
            "Near-full sequence+substrate",
            "Enzyme sequence; substrate SMILES",
            "Protein language-model and molecular graph features with geometric learning",
            "Official public checkpoint model_KcatNet.pt",
            "Twenty-six sequences truncated to 1,000 residues; one invalid Quinate SMILES omitted.",
        ],
        "PreTKcat": [
            "Near-full sequence+substrate",
            "Enzyme sequence; substrate SMILES; temperature",
            "ProtT5, MolGNet, and two normalized temperature features with ExtraTrees",
            "Public-data retraining: 16,249 rows, 1,794 features, 100 trees, seed 42",
            "No fitted kcat regressor was released; 148 temperatures imputed to 30 C; 26 sequences truncated; 26 exact public-corpus pair overlaps were retained.",
        ],
        "SELFprot": [
            "Near-full sequence+substrate",
            "Enzyme sequence; substrate SMILES",
            "Modular protein/molecule transformer with joint kcat head",
            "Released fold-1 checkpoint",
            "Only fold 1 was present in the released weight package; one invalid Quinate SMILES omitted.",
        ],
        "UniKP": [
            "Near-full sequence+substrate",
            "Enzyme sequence; substrate SMILES",
            "Mean-pooled ProtT5 and SMILES Transformer features with ExtraTrees",
            "Official public regressor",
            "Sequences over 1,000 residues use the first and last 500 residues; one invalid Quinate SMILES omitted.",
        ],
        "PMAK": [
            "Reaction-aware",
            "Complete reaction SMILES; enzyme sequence",
            "ProtT5 and RXNFP features with residue-aware attention",
            "Mean of five released reaction-cold fold checkpoints",
            "Requires complete reaction SMILES; 198 benchmark rows were outside the input domain.",
        ],
        "TurNuP": [
            "Reaction-aware",
            "Reactant SMILES; product SMILES; enzyme sequence",
            "DRFP reaction fingerprint and task-specific ESM-1b/ESP representation with gradient boosting",
            "Official public code and checkpoint",
            "Requires complete reactant and product representations; 198 rows were outside the input domain.",
        ],
        "CatPred": [
            "Method-specific subset",
            "Substrate SMILES; enzyme sequence; generated protein-record cache key",
            "Production kcat ensemble using sequence attention/ESM2 and a D-MPNN substrate encoder",
            "Official production kcat checkpoint ensemble",
            "Sixty-four monatomic-proton rows and one invalid Quinate SMILES did not yield predictions; production kcat inference did not use a 3D EGNN.",
        ],
        "KinForm": [
            "Method-specific subset",
            "Enzyme sequence; substrate SMILES; released precomputed protein features",
            "KinForm-L: ESMC, ESM-2 and ProtT5 global/binding-site-weighted features, SMILES Transformer, and ExtraTrees",
            "Released KinForm-L asset bundle",
            "Only 563 rows had all released lookup/features; 415 rows lacked required assets or had invalid SMILES.",
        ],
        "DEKP": [
            "Publicly retrained structure-aware",
            "Enzyme sequence; substrate SMILES; AlphaFold/PDB residue graph",
            "DEKP MetaDecoder with molecular, sequence-CNN, and structure-graph features",
            "DEKP-public-retrained: 13,385 public rows after excluding 16 exact-pair source rows",
            "Final kcat weights were not released; 12,046/1,339 train/validation split; one invalid Quinate SMILES omitted.",
        ],
        "GO-HKP": [
            "Functional-assignment baseline",
            "DeepGO-SE reaction assignments or UniProt GO annotations",
            "GO hierarchy and organism-filtered GO-linked kcat medians",
            "Benchmark functional-assignment baseline",
            "E. coli and yeast use different GO evidence routes; this is not a direct regression model or a guaranteed leakage-free negative control.",
        ],
    }
    for row in range(2, details.max_row + 1):
        method = details.cell(row, 1).value
        if method not in replacements:
            continue
        for col, value in enumerate(replacements[method], start=2):
            details.cell(row, col, value)
        if method == "DEKP":
            details.cell(row, 1, "DEKP-public-retrained")
        elif method == "KinForm":
            details.cell(row, 1, "KinForm-L")


def update_overlap_sheet(ws) -> None:
    for row in range(2, ws.max_row + 1):
        method = ws.cell(row, 1).value
        if method == "PreTKcat":
            ws.cell(row, 2, "Exact enzyme-substrate pair in public fitting corpus (retained)")
            ws.cell(
                row,
                9,
                "The public-data retraining retained these 26 exact pairs; its aggregate score is therefore not a strictly pair-disjoint estimate.",
            )
        elif method == "DEKP":
            ws.cell(row, 1, "DEKP-public-retrained")
            ws.cell(row, 2, "Exact benchmark pair in original public source corpus (excluded before fitting)")
            ws.cell(
                row,
                9,
                "Seventeen benchmark records mapped to 16 source rows that were excluded before retraining; this subgroup is source-corpus proximity, not actual training overlap.",
            )
        elif method == "CatPred":
            level = ws.cell(row, 2).value
            evaluated_subset_counts = {
                "Exact UniProt + reactant SMILES": 3,
                "Same UniProt + substrate component": 24,
                "UniProt overlap": 426,
                "EC overlap": 842,
            }
            if level in evaluated_subset_counts:
                ws.cell(row, 3, evaluated_subset_counts[level])
                ws.cell(row, 4, 913)
                if level == "UniProt overlap":
                    ws.cell(row, 9, "Protein-identity overlap was common within the 913 CatPred-scored records.")
                elif level == "EC overlap":
                    ws.cell(row, 9, "Functional-class overlap covered most of the 913 CatPred-scored records.")


def update_supporting_sheets(wb) -> None:
    update_method_sheets(wb)
    update_wilcoxon_sheet(wb["S8_Wilcoxon"])
    update_overlap_sheet(wb["S14_Overlap_audit"])
    for ws in wb.worksheets:
        if ws.title == "Table0":
            continue
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == "DEKP":
                    cell.value = "DEKP-public-retrained"
                elif cell.value == "KinForm":
                    cell.value = "KinForm-L"


def write_workbook(source: Path, output: Path, table: pd.DataFrame, coverage: dict[str, int]) -> None:
    wb = load_workbook(source)
    if "Table0" in wb.sheetnames:
        del wb["Table0"]
    ws = wb.create_sheet("Table0", 0)
    ws.append(list(table.columns))
    for row in table.itertuples(index=False, name=None):
        ws.append(list(row))
    style_table0(ws, coverage)
    update_index(wb)
    update_supporting_sheets(wb)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def main() -> None:
    args = parse_args()
    table, coverage = load_prediction_matrix()
    validate_against_summary(table, coverage)
    write_workbook(args.input, args.output, table, coverage)
    print(f"Wrote {args.output}")
    print(f"Table0 rows: {len(table)}; columns: {len(table.columns)}")
    print("Coverage: " + ", ".join(f"{method}={n}" for method, n in coverage.items()))


if __name__ == "__main__":
    main()
